from __future__ import annotations

import json
import sqlite3
import unittest

from openpyxl import load_workbook

from services.investmap_rf_monitor_export import build_monitor_export_xlsx


class InvestmapRfMonitorExportTest(unittest.TestCase):
    def setUp(self):
        self.conn = sqlite3.connect(":memory:")
        self.conn.row_factory = sqlite3.Row

        self.conn.executescript(
            """
            CREATE TABLE investmap_rf_card_snapshots (
                id INTEGER PRIMARY KEY,
                global_id INTEGER NOT NULL,
                payload_json TEXT,
                fetched_at_utc TEXT,
                filling_level REAL
            );

            CREATE TABLE portal_analysis_runs (
                id INTEGER PRIMARY KEY,
                created_at TEXT,
                source_label TEXT
            );

            CREATE TABLE portal_analysis_site_snapshots (
                id INTEGER PRIMARY KEY,
                site_id TEXT,
                run_id INTEGER,
                score_percent REAL,
                analysis_status TEXT
            );

            CREATE TABLE investmap_rf_card_manager_assignments (
                global_id INTEGER PRIMARY KEY,
                municipality_raw TEXT,
                manager_name TEXT,
                assignment_source TEXT,
                match_status TEXT
            );

            CREATE TABLE investmap_rf_manager_match_issues (
                id INTEGER PRIMARY KEY,
                global_id INTEGER,
                issue_type TEXT,
                details TEXT,
                is_resolved INTEGER
            );

            CREATE TABLE investmap_fields (
                tech_name TEXT,
                display_name TEXT
            );

            CREATE TABLE investmap_rules (
                source_field TEXT,
                source_value TEXT,
                target_field TEXT,
                recommended_text TEXT
            );
            """
        )

        self.conn.execute(
            """
            INSERT INTO portal_analysis_runs (id, created_at)
            VALUES (1, '2026-08-25T09:00:00Z')
            """
        )

        self._add_card(
            global_id=1001,
            name="Площадка 1001",
            municipality="Муниципалитет 1",
            filling=95,
            manager="Иванов И.И.",
        )
        self._add_card(
            global_id=1002,
            name="Площадка 1002",
            municipality="Муниципалитет 2",
            filling=70,
            manager="Иванов И.И.",
        )
        self._add_card(
            global_id=1003,
            name="Площадка 1003",
            municipality="Муниципалитет 3",
            filling=None,
            manager=None,
            issue_type="unmatched",
        )

        self.conn.commit()

    def tearDown(self):
        self.conn.close()

    def _add_card(
        self,
        *,
        global_id,
        name,
        municipality,
        filling,
        manager,
        issue_type=None,
    ):
        payload = json.dumps(
            {
                "name": name,
                "municipality": municipality,
            },
            ensure_ascii=False,
        )

        self.conn.execute(
            """
            INSERT INTO investmap_rf_card_snapshots (
                id,
                global_id,
                payload_json,
                fetched_at_utc,
                filling_level
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                global_id,
                global_id,
                payload,
                "2026-08-25T10:00:00Z",
                filling,
            ),
        )

        self.conn.execute(
            """
            INSERT INTO portal_analysis_site_snapshots (
                id,
                site_id,
                run_id,
                score_percent,
                analysis_status
            )
            VALUES (?, ?, 1, ?, 'ok')
            """,
            (global_id, str(global_id), 80),
        )

        if manager is not None:
            self.conn.execute(
                """
                INSERT INTO investmap_rf_card_manager_assignments (
                    global_id,
                    municipality_raw,
                    manager_name,
                    assignment_source,
                    match_status
                )
                VALUES (?, ?, ?, 'automatic', 'matched')
                """,
                (global_id, municipality, manager),
            )

        if issue_type is not None:
            self.conn.execute(
                """
                INSERT INTO investmap_rf_manager_match_issues (
                    id,
                    global_id,
                    issue_type,
                    details,
                    is_resolved
                )
                VALUES (?, ?, ?, 'Нет правила', 0)
                """,
                (global_id, global_id, issue_type),
            )

    def test_adds_traffic_light_columns_to_low_score_section(self):
        stream = build_monitor_export_xlsx(self.conn)
        workbook = load_workbook(stream, data_only=True)
        sheet = workbook["Лучшие и худшие площадки"]

        low_score_title_row = next(
            row_number
            for row_number in range(1, sheet.max_row + 1)
            if sheet.cell(row=row_number, column=1).value
            == "Площадки с заполнением ниже 80%"
        )
        header_row = low_score_title_row + 1
        card_row = low_score_title_row + 2

        self.assertEqual(
            [
                sheet.cell(header_row, column).value
                for column in range(1, 7)
            ],
            [
                "Global ID",
                "Средний процент заполнения",
                "Территориальный управляющий",
                "🟢 Повысит % заполняемости",
                "🟡 Может повлиять на % заполняемости",
                "🔴 Влияние на % не подтверждено",
            ],
        )
        self.assertEqual(sheet.cell(card_row, column=1).value, 1002)
        self.assertEqual(sheet.cell(card_row, column=4).value, "—")
        self.assertEqual(sheet.cell(card_row, column=5).value, "—")
        self.assertEqual(sheet.cell(card_row, column=6).value, "—")
        
    def test_builds_expected_sheets_and_summary(self):
        stream = build_monitor_export_xlsx(self.conn)
        workbook = load_workbook(stream, data_only=True)

        self.assertEqual(
            workbook.sheetnames,
            [
                "Техническая информация",
                "Общая информация",
                "Итог по управляющим",
                "Лучшие и худшие площадки",
            ],
        )

        technical = workbook["Техническая информация"]
        self.assertEqual(
            [technical.cell(1, column).value for column in range(1, 11)],
            [
                "Global ID",
                "Наименование",
                "Муниципалитет",
                "Территориальный управляющий",
                "Источник назначения",
                "Статус сопоставления",
                "Открытая проблема",
                "Дата последнего snapshot",
                "Средний процент заполнения",
                "V2-оценка и дата анализа",
            ],
        )
        self.assertEqual(technical["A2"].value, 1001)
        self.assertEqual(technical["B2"].value, "Площадка 1001")
        self.assertEqual(technical["I2"].value, 95)

        summary = workbook["Итог по управляющим"]
        self.assertEqual(summary["A2"].value, "Иванов И.И.")
        self.assertEqual(summary["B2"].value, 2)
        self.assertEqual(summary["C2"].value, 82.5)

        self.assertEqual(summary["A4"].value, "Итого")
        self.assertEqual(summary["B4"].value, 3)
        self.assertEqual(summary["C4"].value, 82.5)

    def test_places_low_and_missing_cards_in_rankings(self):
        stream = build_monitor_export_xlsx(self.conn)
        workbook = load_workbook(stream, data_only=True)
        sheet = workbook["Лучшие и худшие площадки"]
        titles = {
            sheet.cell(row=row_number, column=1).value: row_number
            for row_number in range(1, sheet.max_row + 1)
            if sheet.cell(row=row_number, column=1).value
            in {
                "Лучшие 10 площадок",
                "Площадки с заполнением ниже 80%",
                "Карточки без рассчитанного процента",
            }
        }

        self.assertEqual(titles["Лучшие 10 площадок"], 1)
        self.assertEqual(
            sheet.cell(
                row=titles["Лучшие 10 площадок"] + 2,
                column=1,
            ).value,
            1001,
        )
        self.assertEqual(
            sheet.cell(
                row=titles["Площадки с заполнением ниже 80%"] + 2,
                column=1,
            ).value,
            1002,
        )
        self.assertEqual(
            sheet.cell(
                row=titles["Карточки без рассчитанного процента"] + 2,
                column=1,
            ).value,
            1003,
        )
        self.assertEqual(
            sheet.cell(
                row=titles["Карточки без рассчитанного процента"] + 2,
                column=2,
            ).value,
            "—",
        )
