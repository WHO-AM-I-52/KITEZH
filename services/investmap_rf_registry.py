from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


FREE_STATUS = "Свободна"
EVENT_ACTIVATED = "activated"
EVENT_REACTIVATED = "reactivated"
EVENT_DEACTIVATED_STATUS_CHANGED = "deactivated_status_changed"
EVENT_DEACTIVATED_API_NOT_FOUND = "deactivated_api_not_found"


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().casefold().split())


def _normalize_status(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _parse_global_id(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, int):
        return value if value > 0 else None

    if isinstance(value, float):
        if value.is_integer() and value > 0:
            return int(value)
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        parsed = int(text)
    except ValueError:
        return None

    return parsed if parsed > 0 else None


def _append_event(
    conn,
    *,
    global_id: int,
    event_type: str,
    previous_status: str | None,
    current_status: str,
    source_filename: str,
    occurred_at_utc: str,
    reason: str | None = None,
    changed_by_user_id: int | None = None,
) -> None:
    conn.execute(
        """
        INSERT INTO investmap_rf_monitor_registry_events (
            global_id,
            event_type,
            previous_status,
            current_status,
            source_filename,
            occurred_at_utc,
            reason,
            changed_by_user_id
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            global_id,
            event_type,
            previous_status,
            current_status,
            source_filename,
            occurred_at_utc,
            reason,
            changed_by_user_id,
        ),
    )


def import_monitored_cards_xlsx(
    conn,
    xlsx_bytes: bytes,
    source_filename: str,
) -> dict[str, Any]:
    """
    Импортирует статусы площадок из XLSX в реестр мониторинга.

    Функция не вызывает commit() и не закрывает переданное соединение conn.
    """
    if load_workbook is None:
        raise RuntimeError(
            "Для импорта XLSX требуется пакет openpyxl. "
            "Установите его командой: pip install openpyxl"
        )

    if not isinstance(xlsx_bytes, (bytes, bytearray)) or not xlsx_bytes:
        raise ValueError("Файл XLSX пустой или передан в недопустимом формате.")

    normalized_filename = str(source_filename or "").strip()
    if not normalized_filename:
        raise ValueError("Не задано имя исходного файла.")

    try:
        workbook = load_workbook(
            filename=BytesIO(xlsx_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValueError("Не удалось прочитать XLSX-файл.") from exc

    report: dict[str, Any] = {
        "source_filename": normalized_filename,
        "sheet_name": None,
        "rows_total": 0,
        "rows_with_global_id": 0,
        "free_rows": 0,
        "non_free_rows": 0,
        "invalid_rows": 0,
        "duplicate_rows": 0,
        "added": 0,
        "reactivated": 0,
        "deactivated": 0,
        "unchanged_active": 0,
        "unchanged_inactive": 0,
        "errors": [],
    }

    try:
        worksheet = workbook.active
        report["sheet_name"] = worksheet.title

        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValueError("Первый лист XLSX не содержит строк.") from exc

        header_positions: dict[str, int] = {}
        for index, header in enumerate(raw_headers):
            normalized = _normalize_header(header)
            if normalized and normalized not in header_positions:
                header_positions[normalized] = index

        global_id_index = header_positions.get("global_id")
        status_index = header_positions.get("статус площадки")

        missing_headers: list[str] = []
        if global_id_index is None:
            missing_headers.append("global_id")
        if status_index is None:
            missing_headers.append("Статус площадки")

        if missing_headers:
            missing_text = ", ".join(missing_headers)
            raise ValueError(
                f"В XLSX не найдены обязательные столбцы: {missing_text}."
            )

        seen_global_ids: set[int] = set()
        imported_at_utc = _utc_now()

        for row_number, row in enumerate(rows, start=2):
            report["rows_total"] += 1

            global_id_value = (
                row[global_id_index] if global_id_index < len(row) else None
            )
            status_value = row[status_index] if status_index < len(row) else None

            global_id = _parse_global_id(global_id_value)
            if global_id is None:
                report["invalid_rows"] += 1
                report["errors"].append(
                    {
                        "row": row_number,
                        "reason": "Некорректный или отсутствующий global_id.",
                    }
                )
                continue

            report["rows_with_global_id"] += 1

            if global_id in seen_global_ids:
                report["duplicate_rows"] += 1
                report["errors"].append(
                    {
                        "row": row_number,
                        "global_id": global_id,
                        "reason": "Повторный global_id в исходном файле.",
                    }
                )
                continue

            seen_global_ids.add(global_id)

            status = _normalize_status(status_value)
            if not status:
                report["invalid_rows"] += 1
                report["errors"].append(
                    {
                        "row": row_number,
                        "global_id": global_id,
                        "reason": "Не указан статус площадки.",
                    }
                )
                continue

            row_exists = conn.execute(
                """
                SELECT is_active, last_source_status
                FROM investmap_rf_monitored_cards
                WHERE global_id = ?
                """,
                (global_id,),
            ).fetchone()

            if status == FREE_STATUS:
                report["free_rows"] += 1

                if row_exists is None:
                    conn.execute(
                        """
                        INSERT INTO investmap_rf_monitored_cards (
                            global_id,
                            is_active,
                            source_filename,
                            imported_at_utc,
                            last_seen_import_at_utc,
                            last_source_status
                        )
                        VALUES (?, 1, ?, ?, ?, ?)
                        """,
                        (
                            global_id,
                            normalized_filename,
                            imported_at_utc,
                            imported_at_utc,
                            status,
                        ),
                    )
                    _append_event(
                        conn,
                        global_id=global_id,
                        event_type=EVENT_ACTIVATED,
                        previous_status=None,
                        current_status=status,
                        source_filename=normalized_filename,
                        occurred_at_utc=imported_at_utc,
                    )
                    report["added"] += 1
                    continue

                previous_is_active = int(row_exists["is_active"])
                previous_status = row_exists["last_source_status"]

                conn.execute(
                    """
                    UPDATE investmap_rf_monitored_cards
                    SET
                        is_active = 1,
                        source_filename = ?,
                        last_seen_import_at_utc = ?,
                        last_source_status = ?
                    WHERE global_id = ?
                    """,
                    (
                        normalized_filename,
                        imported_at_utc,
                        status,
                        global_id,
                    ),
                )

                if previous_is_active == 0:
                    _append_event(
                        conn,
                        global_id=global_id,
                        event_type=EVENT_REACTIVATED,
                        previous_status=previous_status,
                        current_status=status,
                        source_filename=normalized_filename,
                        occurred_at_utc=imported_at_utc,
                    )
                    report["reactivated"] += 1
                else:
                    report["unchanged_active"] += 1

                continue

            report["non_free_rows"] += 1

            if row_exists is None:
                continue

            previous_is_active = int(row_exists["is_active"])
            previous_status = row_exists["last_source_status"]

            conn.execute(
                """
                UPDATE investmap_rf_monitored_cards
                SET
                    source_filename = ?,
                    last_seen_import_at_utc = ?,
                    last_source_status = ?
                WHERE global_id = ?
                """,
                (
                    normalized_filename,
                    imported_at_utc,
                    status,
                    global_id,
                ),
            )

            if previous_is_active == 1:
                conn.execute(
                    """
                    UPDATE investmap_rf_monitored_cards
                    SET is_active = 0
                    WHERE global_id = ?
                    """,
                    (global_id,),
                )
                _append_event(
                    conn,
                    global_id=global_id,
                    event_type=EVENT_DEACTIVATED_STATUS_CHANGED,
                    previous_status=previous_status,
                    current_status=status,
                    source_filename=normalized_filename,
                    occurred_at_utc=imported_at_utc,
                )
                report["deactivated"] += 1
            else:
                report["unchanged_inactive"] += 1

    finally:
        workbook.close()

    return report
