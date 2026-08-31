"""Read-only Excel export for the InvestMap RF monitor."""

from __future__ import annotations

from collections import defaultdict
from html import unescape
from io import BytesIO
import json
import re
import sqlite3
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from portal_analysis.api_snapshot_normalizer import api_snapshot_to_portal_row
from portal_analysis.portal_checker import calc_portal_score_v2


_TECHNICAL_HEADERS = [
    "Global ID",
    "Наименование",
    "Муниципалитет",
    "Территориальный управляющий",
    "Источник назначения",
    "Статус сопоставления",
    "Открытая проблема",
    "Дата последнего snapshot",
    "Средний процент заполнения",
    "V2-оценка и дата анализа",
]

_API_FIELD_MASK = {
    "id": "ID площадки в API",
    "siteStatus": "Статус площадки",
    "siteName": "Название площадки",
    "regions": "Регион",
    "municipality": "Муниципальное образование",
    "adressObject": "Адрес объекта",
    "nearestCity": "Ближайший город",
    "formatSites": "Формат площадки",
    "typeSites": "Тип площадки",
    "distanceFromRoad": "Удаленность от автомобильной дороги (метров)",
    "presenceOfEncumbrancesDictTitles": "Наличие обременений",
    "addInfoAboutEncumbrances": "Дополнительная информация об обременении",
    "isPriority": "Приоритетная площадка",
    "formOwnerships": "Форма собственности объекта",
    "transactionForms": "Форма сделки",
    "costObject": "Стоимость объекта, руб. (покупки или месячной аренды)",
    "costPerHa": "Стоимость, руб./год за га",
    "costPerSqM": "Стоимость, руб./год за кв.м.",
    "maxMinRentalPeriod": "Min и max сроки аренды (если применимо), лет",
    "maxMinRentalYear": "Min и max сроки аренды, лет",
    "procedureDeterminingCostStr": "Порядок определения стоимости",
    "hazardClassObjects": "Класс опасности объекта",
    "characteristicsCapitalBuildings": (
        "Характеристики расположенных объектов капитального строительства"
    ),
    "landArea": "Свободная площадь ЗУ, га",
    "cadastralLandNumber": "Кадастровый номер ЗУ",
    "typesAuthorizedUses": "Варианты разрешенного использования",
    "landSurveying": "Межевание ЗУ",
    "landCategory": "Категория земель",
    "areaPropertyComplex": (
        "Свободная площадь здания, сооружения, помещения, кв. м"
    ),
    "totalSiteArea": "Площадь объекта, кв. м",
    "cadastralPropertyComplexNumber": (
        "Кадастровый номер здания, сооружения, помещения"
    ),
    "buildingSpecifications": (
        "Технические характеристики здания, сооружения, помещения"
    ),
    "nameOwner": "Наименование собственника / администратора объекта",
    "inn": "ИНН собственника",
    "contactPerson": "Контактное лицо",
    "contactPhoneNumber": "Телефон контактного лица, e-mail",
    "websiteContactPerson": "Сайт",
    "notes": "Примечание",
    "waterSupply.availability": "Водоснабжение Наличие (Да/Нет)",
    "waterSupply.connectionFeeMin": (
        "Водоснабжение Плата за подключение (минимальная), руб."
    ),
    "waterSupply.connectionFeeMax": (
        "Водоснабжение Плата за подключение (максимальная), руб."
    ),
    "waterSupply.tariffDrinkingWater": (
        "Водоснабжение Тариф на питьевую воду, руб./куб.м"
    ),
    "waterSupply.tariffProcessWater": (
        "Водоснабжение Тариф на техническую воду, руб./куб.м"
    ),
    "waterSupply.tariffTransportation": (
        "Водоснабжение Тариф на транспортировку воды, руб./куб.м"
    ),
    "waterSupply.availableCapacity": (
        "Объекты водоснабжения Максимально допустимая мощность, куб. м/ч"
    ),
    "waterSupply.freePower": "Объекты водоснабжения Свободная мощность, куб.м/ч",
    "waterSupply.bandwidth": "Сети водоснабжения Пропускная способность, куб. м/ч",
    "waterSupply.otherFreePower": "Объекты водоснабжения Иные характеристики",
    "waterDisposal.availability": "Водоотведение Наличие (Да/Нет)",
    "waterDisposal.connectionFeeMin": (
        "Водоотведение Плата за подключение (минимальная), руб."
    ),
    "waterDisposal.connectionFeeMax": (
        "Водоотведение Плата за подключение (максимальная), руб."
    ),
    "waterDisposal.tariffWaterDisposal": (
        "Водоотведение Тариф на водоотведение, руб./куб.м"
    ),
    "waterDisposal.bandwidth": (
        "Сети водоотведения Пропускная способность, куб. м/ч"
    ),
    "waterDisposal.otherFreePower": "Объекты водоотведения Иные характеристики",
    "gasSupply.availability": "Газоснабжение Наличие (Да/Нет)",
    "gasSupply.connectionFeeMin": (
        "Газоснабжение Плата за подключение (минимальная), руб."
    ),
    "gasSupply.connectionFeeMax": (
        "Газоснабжение Плата за подключение (максимальная), руб."
    ),
    "gasSupply.basicWholesalePriceGas": (
        "Газоснабжение Базовая оптовая цена на газ, руб./1000 куб. м"
    ),
    "gasSupply.tariffTransportation": (
        "Газоснабжение Тариф на услуги по транспортировке газа, руб./1000 куб. м"
    ),
    "gasSupply.consumerGroupByVolumeDictTitles": (
        "Газоснабжение Группа потребителей (по объему потребления)"
    ),
    "gasSupply.specialSurchargeToTransportationTariff": (
        "Газоснабжение Спец. надбавка к тарифу на транспортировку газа, "
        "руб./1000 куб. м"
    ),
    "gasSupply.feeSupplyAndDistributionServices": (
        "Газоснабжение Плата за снабженческо-сбытовые услуги, руб./куб. м"
    ),
    "gasSupply.availableCapacity": (
        "Объекты газоснабжения Величина максимального расхода газа "
        "(мощности) газоиспользующего оборудования, куб. м./ч"
    ),
    "gasSupply.freePower": "Объекты газоснабжения Свободная мощность, куб. м/ч",
    "gasSupply.tariffConsumption": (
        "Газоснабжение Тариф на потребление, руб./куб. м"
    ),
    "gasSupply.otherFreePower": "Объекты газоснабжения Иные характеристики",
    "powerSupply.availability": "Электроснабжение Наличие (Да/Нет)",
    "powerSupply.connectionFeeMin": (
        "Электроснабжение Плата за подключение (минимальная), руб."
    ),
    "powerSupply.connectionFeeMax": (
        "Электроснабжение Плата за подключение (максимальная), руб."
    ),
    "powerSupply.priceCategoryDictTitles": "Электроснабжение Ценовая категория (ЦК)",
    "powerSupply.tariffTransportation": (
        "Объекты электроснабжения Тариф на услуги по передаче "
        "электрической энергии, руб./МВт.ч"
    ),
    "powerSupply.salesSurchargeGuaranteedSupplier": (
        "Электроснабжение Сбытовая надбавка гарантирующего поставщика, руб./кВт.ч"
    ),
    "powerSupply.tariffOperationalDispatchManagement": (
        "Объекты электроснабжения Тариф на услуги по оперативно-диспетчерскому "
        "управлению, руб./кВт.ч"
    ),
    "powerSupply.availableCapacity": (
        "Объекты электроснабжения Максимальная мощность, МВт/ч"
    ),
    "powerSupply.freePower": "Объекты электроснабжения Свободная мощность, МВт/ч",
    "powerSupply.powerCost": "Электроснабжение Стоимость мощности, руб./кВт.ч",
    "powerSupply.tariffConsumption": (
        "Электроснабжение Тариф на потребление, руб./МВт.ч"
    ),
    "powerSupply.otherFreePower": "Объекты электроснабжения Иные характеристики",
    "heatSupply.availability": "Теплоснабжение Наличие (Да/Нет)",
    "heatSupply.connectionFeeMin": (
        "Теплоснабжение Плата за подключение (минимальная), руб."
    ),
    "heatSupply.connectionFeeMax": (
        "Теплоснабжение Плата за подключение (максимальная), руб."
    ),
    "heatSupply.tariffThermalEnergyPower": (
        "Теплоснабжение Тариф на тепловую энергию (мощность), руб./Гкал"
    ),
    "heatSupply.tariffHeatCarrier": (
        "Теплоснабжение Тариф на теплоноситель, руб./Гкал"
    ),
    "heatSupply.tariffTransportation": (
        "Теплоснабжение Тариф на услуги по передаче тепловой энергии, руб./Гкал"
    ),
    "heatSupply.feeHeatCapacityMaintenanceService": (
        "Теплоснабжение Плата за услуги по поддержанию резервной тепловой "
        "мощности, руб./Гкал"
    ),
    "heatSupply.tariffConsumption": "Теплоснабжение Тариф на потребление, руб./Гкал",
    "heatSupply.otherFreePower": "Объекты теплоснабжения Иные характеристики",
    "mswRemoval.availability": "Вывоз ТКО Наличие (Да/Нет)",
    "mswRemoval.tariffMSWRemovalDictTitles": "Вид тарифа на вывоз ТКО",
    "mswRemoval.tariffInCubicMetres": "Тариф на вывоз ТКО, руб./куб. м",
    "mswRemoval.tariff": "Тариф на вывоз ТКО, руб./мес",
    "mswRemoval.otherFree": "Объекты ТКО Иные характеристики",
    "accessRoadsAvailability": "Наличие подъездных путей (Да/Нет)",
    "railwayAvailability": "Наличие ж/д (Да/Нет)",
    "truckParkingAvailability": "Наличие парковки грузового транспорта",
    "otherInformationSite": "Иные сведения",
    "descriptionApplicationProcedure": "Описание процедуры подачи заявки",
    "listOfDocumentsForApplication": "Перечень документов, необходимых для подачи заявки",
    "emailAddressForApplying": "Адрес эл. почты для подачи заявки",
    "linkToApplicationForm": "Ссылка на форму подачи заявки",
    "economicActivitiesForImplementations": (
        "Перечень видов экономической деятельности, возможных к реализации "
        "на площадке"
    ),
    "urbanPlanCharacteristicsAndLimits": "Градостроительные характеристики и ограничения",
    "territorialPlanDocumentsFile": "Документы территориального планирования",
    "photos": "Фотографии объекта",
    "documents": "Документы по объекту",
    "isSupportMeasureSite": "Наличие МАИП",
    "preferentialTreatmentSites": "Преференциальный режим",
    "preferentialBusinessLink.name": "Наименование объекта преференциального режима",
    "preferentialBusinessLink.incomeTax": "Описание льготы: налог на прибыль",
    "preferentialBusinessLink.propertyTax": "Описание льготы: налог на имущество",
    "preferentialBusinessLink.landTax": "Описание льготы: земельный налог",
    "preferentialBusinessLink.transportTax": "Описание льготы: транспортный налог",
    "preferentialBusinessLink.insurancePremiums": "Описание льготы: страховые взносы",
    "preferentialBusinessLink.otherBenefits": "Описание льготы: прочее",
    "businessEnvironmentPreferentialLink.name": "Наименование преференциальной среды",
    "businessEnvironmentPrivilegedLink.name": "Наименование льготной среды",
    "supportInfrastructureLink.name": "Наименование инфраструктуры поддержки",
    "supportInfrastructureObjects": "Объект инфраструктуры поддержки",
    "coordinate.latitude": "Широта объекта в координатах WGS-84",
    "coordinate.longitude": "Долгота объекта в координатах WGS-84",
}

_OVERVIEW_HEADERS = [
    "Global ID",
    "Наименование",
    "Муниципалитет",
    "Территориальный управляющий",
    "Средний процент заполнения",
]

_MANAGER_HEADERS = [
    "Территориальный управляющий",
    "Количество карточек",
    "Средний процент заполнения",
]

_RANKING_HEADERS = [
    "Global ID",
    "Средний процент заполнения",
    "Территориальный управляющий",
    "🟢 Повысит % заполняемости",
    "🟡 Может повлиять на % заполняемости",
    "🔴 Влияние на % не подтверждено",
]

_STATUS_LABELS = {
    "matched": "Сопоставлено",
    "manual": "Назначено вручную",
    "unmatched": "Не найдено",
    "ambiguous": "Неоднозначно",
}

_SOURCE_LABELS = {
    "auto": "Автоматическое",
    "automatic": "Автоматическое",
    "api_contact_person": "Контактное лицо API",
    "manual": "Ручное",
}

_ISSUE_LABELS = {
    "unmatched": "Не найдено правило",
    "ambiguous": "Несколько правил",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)

_HTML_BREAK_PATTERN = re.compile(r"<\s*br\s*/?\s*>", re.IGNORECASE)
_HTML_BLOCK_PATTERN = re.compile(
    r"</?\s*(?:p|div|li|ul|ol|strong|em|b|i|u)[^>]*>",
    re.IGNORECASE,
)
_HTML_TAG_PATTERN = re.compile(r"<[^>]+>")
_S3_FILE_PATTERN = re.compile(
    r"S3:file_storage/[0-9a-fA-F-]{36}_"
)


def _clean_text(value: str) -> str:
    value = unescape(value)
    value = _HTML_BREAK_PATTERN.sub("\n", value)
    value = _HTML_BLOCK_PATTERN.sub("\n", value)
    value = _HTML_TAG_PATTERN.sub("", value)
    value = value.replace("\xa0", " ")
    value = re.sub(r"[ \t]+\n", "\n", value)
    value = re.sub(r"\n[ \t]+", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    value = re.sub(r"[ \t]{2,}", " ", value)
    return value.strip()


def _clean_s3_file_references(value: str) -> str:
    return _S3_FILE_PATTERN.sub("", value)


def _first_text(value: Any) -> str | None:
    if isinstance(value, str):
        value = _clean_text(value)
        return value or None
    return None


def _payload_value(payload: Any, *keys: str) -> str | None:
    if not isinstance(payload, dict):
        return None

    for key in keys:
        value = _first_text(payload.get(key))
        if value:
            return value

    return None


def _flatten_payload(
    value: Any,
    path: str = "",
) -> dict[str, Any]:
    """Возвращает плоское представление JSON; списки остаются одной ячейкой."""
    if isinstance(value, dict):
        flattened: dict[str, Any] = {}

        for key in sorted(value, key=str):
            key_path = f"{path}.{key}" if path else str(key)
            flattened.update(_flatten_payload(value[key], key_path))

        return flattened

    if isinstance(value, list):
        return {path: value}

    return {path: value}


def _format_list_item(value: Any) -> str:
    if isinstance(value, dict):
        for key in ("description", "name", "title", "value", "code", "key", "id"):
            text = _first_text(value.get(key))
            if text:
                return text
        return _clean_text(
            json.dumps(
                value,
                ensure_ascii=False,
                separators=(",", ":"),
                default=str,
            )
        )

    if isinstance(value, bool):
        return "Да" if value else "Нет"

    if value is None:
        return ""

    if isinstance(value, str):
        return _clean_s3_file_references(_clean_text(value))

    return str(value).strip()


def _payload_cell_value(value: Any) -> Any:
    """Приводит API-значение к читаемому и поддерживаемому Excel-значению."""
    if value is None:
        return None

    if isinstance(value, bool):
        return "Да" if value else "Нет"

    if isinstance(value, list):
        values = [_format_list_item(item) for item in value]
        return "; ".join(item for item in values if item) or None

    if isinstance(value, dict):
        return _clean_text(
            json.dumps(
                value,
                ensure_ascii=False,
                separators=(",", ":"),
                default=str,
            )
        )

    if isinstance(value, str):
        value = _clean_text(value)
        value = _clean_s3_file_references(value)
        return value or None

    return value

_RED_RECOMMENDATION_FIELDS = {
    "Преференциальный режим",
    "Наименование объекта преференциального режима",
    "Наименование объекта инфраструктуры поддержки",
}

_YELLOW_RECOMMENDATION_FIELDS = {
    "Стоимость, руб./год за кв. м",
    "min и max сроки аренды (если применимо), лет",
    "Свободная площадь ЗУ, га",
    "Варианты разрешенного использования",
    "Межевание земельного участка",
    "Категория земель",
}

_RENTAL_PERIOD_RECOMMENDATION = (
    "Укажите минимальный и максимальный срок аренды в годах, например: 1–49."
)


def _format_recommendation(field: str, hint: str | None) -> str:
    if field == "min и max сроки аренды (если применимо), лет":
        return _RENTAL_PERIOD_RECOMMENDATION
    if field == "Описание процедуры подачи заявки":
        return (
            "Опишите процедуру подачи заявки: порядок обращения, "
            "способ подачи заявки и порядок её рассмотрения."
        )
    if hint:
        return hint
    return f'Заполните поле «{field}».'


def _partial_recommendation(field: str, hint: str | None) -> str:
    """Формирует рекомендацию для частично заполненного поля V2."""
    if hint:
        return hint

    if field.endswith("— Наличие"):
        return (
            f"Уточните значение «{field}». Если коммуникация фактически "
            "доступна, укажите «Да» и внесите подтверждённые параметры "
            "подключения: мощность, тариф, расстояние до точки подключения "
            "или иные сведения."
        )

    return (
        f"Уточните значение «{field}» и добавьте конкретные подтверждённые "
        "параметры вместо предварительной формулировки."
    )


def _traffic_light_recommendations(
    diagnostics: dict[str, Any] | None,
) -> tuple[str, str, str]:
    """Распределяет missing- и partial-поля V2 по колонкам светофора."""
    if diagnostics is None:
        return "—", "—", "—"

    green: list[str] = []
    yellow: list[str] = []
    red: list[str] = []

    for entry in diagnostics.get("missing", []):
        field = entry.get("field")
        if not field:
            continue

        recommendation = _format_recommendation(field, entry.get("hint"))
        if field in _RED_RECOMMENDATION_FIELDS:
            red.append(recommendation)
        elif field in _YELLOW_RECOMMENDATION_FIELDS:
            yellow.append(recommendation)
        else:
            green.append(recommendation)

    for entry in diagnostics.get("partial", []):
        field = entry.get("field")
        if not field:
            continue

        yellow.append(
            _partial_recommendation(field, entry.get("hint"))
        )

    return (
        "\n".join(green) or "—",
        "\n".join(yellow) or "—",
        "\n".join(red) or "—",
    )


def _read_export_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        WITH latest_snapshot_ids AS (
            SELECT global_id, MAX(id) AS snapshot_id
            FROM investmap_rf_card_snapshots
            GROUP BY global_id
        ),
        latest_snapshots AS (
            SELECT
                snapshots.id AS snapshot_id,
                snapshots.global_id,
                snapshots.payload_json,
                snapshots.fetched_at_utc,
                snapshots.filling_level
            FROM investmap_rf_card_snapshots AS snapshots
            INNER JOIN latest_snapshot_ids
                ON latest_snapshot_ids.snapshot_id = snapshots.id
        ),
        latest_v2_snapshot_ids AS (
            SELECT snapshots.site_id, MAX(snapshots.run_id) AS run_id
            FROM portal_analysis_site_snapshots AS snapshots
            INNER JOIN portal_analysis_runs AS runs
                ON runs.id = snapshots.run_id
            WHERE COALESCE(runs.source_label, '') != 'manual_reanalyze_latest_snapshots'
            GROUP BY snapshots.site_id
        ),
        latest_v2 AS (
            SELECT
                snapshots.site_id,
                snapshots.score_percent,
                snapshots.analysis_status,
                runs.created_at AS analyzed_at_utc
            FROM portal_analysis_site_snapshots AS snapshots
            INNER JOIN latest_v2_snapshot_ids
                ON latest_v2_snapshot_ids.site_id = snapshots.site_id
               AND latest_v2_snapshot_ids.run_id = snapshots.run_id
            INNER JOIN portal_analysis_runs AS runs
                ON runs.id = snapshots.run_id
        ),
        assignments AS (
            SELECT
                global_id,
                municipality_raw,
                manager_name,
                assignment_source,
                match_status
            FROM investmap_rf_card_manager_assignments
        ),
        latest_open_issue_ids AS (
            SELECT global_id, MAX(id) AS issue_id
            FROM investmap_rf_manager_match_issues
            WHERE is_resolved = 0
            GROUP BY global_id
        ),
        open_issues AS (
            SELECT
                issues.global_id,
                issues.issue_type,
                issues.details
            FROM investmap_rf_manager_match_issues AS issues
            INNER JOIN latest_open_issue_ids
                ON latest_open_issue_ids.issue_id = issues.id
        )
        SELECT
            latest.global_id,
            latest.payload_json,
            latest.fetched_at_utc,
            latest.filling_level,
            latest_v2.score_percent AS v2_score,
            latest_v2.analysis_status AS v2_analysis_status,
            latest_v2.analyzed_at_utc AS v2_analyzed_at_utc,
            assignments.municipality_raw,
            assignments.manager_name,
            assignments.assignment_source,
            assignments.match_status,
            open_issues.issue_type AS open_issue_type,
            open_issues.details AS open_issue_details
        FROM latest_snapshots AS latest
        LEFT JOIN latest_v2
            ON latest_v2.site_id = CAST(latest.global_id AS TEXT)
        LEFT JOIN assignments
            ON assignments.global_id = latest.global_id
        LEFT JOIN open_issues
            ON open_issues.global_id = latest.global_id
        ORDER BY latest.global_id ASC
        """
    ).fetchall()

    export_rows: list[dict[str, Any]] = []

    for row in rows:
        payload: Any = None

        try:
            payload = json.loads(row["payload_json"])
        except (TypeError, ValueError):
            payload = None

        municipality = (
            _first_text(row["municipality_raw"])
            or _payload_value(payload, "municipality", "municipality_name")
        )

        name = _payload_value(
            payload,
            "name",
            "title",
            "object_name",
            "site_name",
            "siteName",
            "investment_site_name",
        )

        v2_value = None
        if row["v2_analysis_status"] == "ok" and row["v2_score"] is not None:
            v2_value = f"{row['v2_score']}%"
            if row["v2_analyzed_at_utc"]:
                v2_value += f" ({row['v2_analyzed_at_utc']})"

        issue = _ISSUE_LABELS.get(row["open_issue_type"])
        details = _first_text(row["open_issue_details"])
        if issue and details:
            issue = f"{issue}: {details}"
        elif details:
            issue = details
        diagnostics = None
        if payload and row["v2_analysis_status"] == "ok":
            diagnostics = calc_portal_score_v2(
                api_snapshot_to_portal_row(payload),
                conn,
            )

        green_recommendations, yellow_recommendations, red_recommendations = (
            _traffic_light_recommendations(diagnostics)
        )
        export_rows.append(
            {
                "global_id": row["global_id"],
                "name": name,
                "municipality": municipality,
                "manager_name": _first_text(row["manager_name"]),
                "assignment_source": _SOURCE_LABELS.get(
                    row["assignment_source"],
                    _first_text(row["assignment_source"]),
                ),
                "match_status": _STATUS_LABELS.get(
                    row["match_status"],
                    _first_text(row["match_status"]),
                ),
                "open_issue": issue,
                "fetched_at_utc": row["fetched_at_utc"],
                "filling_level": row["filling_level"],
                "v2_value": v2_value,
                "green_recommendations": green_recommendations,
                "yellow_recommendations": yellow_recommendations,
                "red_recommendations": red_recommendations,
                "api_fields": _flatten_payload(payload) if payload else {},
            }
        )

    return export_rows


def _style_header(sheet, row_number: int, headers: list[str]) -> None:
    for column_number, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_number, column=column_number, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.freeze_panes = f"A{row_number + 1}"


def _apply_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _apply_api_widths(sheet, start_column: int, api_headers: list[str]) -> None:
    for offset, field_path in enumerate(api_headers):
        header = _API_FIELD_MASK.get(field_path, field_path)
        width = min(max(len(header) + 2, 18), 42)
        sheet.column_dimensions[
            get_column_letter(start_column + offset)
        ].width = width


def _technical_api_headers(rows: list[dict[str, Any]]) -> list[str]:
    available_fields = {
        field_path
        for row in rows
        for field_path in row["api_fields"]
        if field_path
    }
    mapped_fields = [
        field_path
        for field_path in _API_FIELD_MASK
        if field_path in available_fields
    ]
    unmapped_fields = sorted(
        available_fields.difference(_API_FIELD_MASK),
        key=str.casefold,
    )
    return [*mapped_fields, *unmapped_fields]


def _append_technical_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.active
    sheet.title = "Техническая информация"

    api_headers = _technical_api_headers(rows)
    visible_api_headers = [
        _API_FIELD_MASK.get(field_path, field_path)
        for field_path in api_headers
    ]
    headers = [*_TECHNICAL_HEADERS, *visible_api_headers]

    _style_header(sheet, 1, headers)
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    )

    for row in rows:
        technical_values = [
            row["global_id"],
            row["name"] or "—",
            row["municipality"] or "—",
            row["manager_name"] or "—",
            row["assignment_source"] or "—",
            row["match_status"] or "Нет данных",
            row["open_issue"] or "—",
            row["fetched_at_utc"] or "—",
            (
                float(row["filling_level"]) / 100
                if row["filling_level"] is not None
                else None
            ),
            row["v2_value"] or "—",
        ]
        api_values = [
            _payload_cell_value(row["api_fields"].get(field_path))
            for field_path in api_headers
        ]
        sheet.append([*technical_values, *api_values])

    for cell in sheet["I"][1:]:
        cell.number_format = "0.0%"

    _apply_widths(sheet, [14, 36, 34, 32, 22, 24, 52, 25, 27, 30])
    _apply_api_widths(sheet, len(_TECHNICAL_HEADERS) + 1, api_headers)


def _append_overview_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Общая информация")
    _style_header(sheet, 1, _OVERVIEW_HEADERS)
    sheet.auto_filter.ref = f"A1:E{len(rows) + 1}"

    for row in rows:
        sheet.append(
            [
                row["global_id"],
                row["name"] or "—",
                row["municipality"] or "—",
                row["manager_name"] or "—",
                (
                    float(row["filling_level"]) / 100
                    if row["filling_level"] is not None
                    else None
                ),
            ]
        )

    for cell in sheet["E"][1:]:
        cell.number_format = "0.0%"

    _apply_widths(sheet, [14, 40, 36, 32, 27])


def _append_manager_summary_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
) -> None:
    sheet = workbook.create_sheet("Итог по управляющим")
    _style_header(sheet, 1, _MANAGER_HEADERS)

    grouped: dict[str, list[float]] = defaultdict(list)
    counts: dict[str, int] = defaultdict(int)

    for row in rows:
        manager = row["manager_name"] or "Не назначен"
        counts[manager] += 1

        if row["filling_level"] is not None:
            grouped[manager].append(float(row["filling_level"]) / 100)

    for manager in sorted(counts, key=lambda item: item.casefold()):
        values = grouped[manager]
        average = sum(values) / len(values) if values else None
        sheet.append([manager, counts[manager], average])

    all_values = [
        float(row["filling_level"]) / 100
        for row in rows
        if row["filling_level"] is not None
    ]
    total_average = sum(all_values) / len(all_values) if all_values else None

    total_row = sheet.max_row + 1
    sheet.append(["Итого", len(rows), total_average])

    for cell in sheet[total_row]:
        cell.fill = _TOTAL_FILL
        cell.font = _BOLD_FONT

    for cell in sheet["C"][1:]:
        cell.number_format = "0.0%"

    _apply_widths(sheet, [36, 22, 30])


def _append_rankings_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Лучшие и худшие площадки")

    calculated = [
        row for row in rows if row["filling_level"] is not None
    ]
    calculated.sort(key=lambda row: (-float(row["filling_level"]), row["global_id"]))
    top_ten = calculated[:10]

    below_eighty = [
        row
        for row in rows
        if row["filling_level"] is not None and row["filling_level"] < 80
    ]
    below_eighty.sort(key=lambda row: (float(row["filling_level"]), row["global_id"]))

    without_filling = [
        row for row in rows if row["filling_level"] is None
    ]
    without_filling.sort(key=lambda row: row["global_id"])

    def append_section(
        title: str,
        items: list[dict[str, Any]],
        start_row: int,
        *,
        include_recommendations: bool = False,
    ) -> int:
        title_cell = sheet.cell(row=start_row, column=1, value=title)
        title_cell.fill = _SECTION_FILL
        title_cell.font = _BOLD_FONT

        headers = _RANKING_HEADERS if include_recommendations else _RANKING_HEADERS[:3]
        _style_header(sheet, start_row + 1, headers)

        row_number = start_row + 2
        for item in items:
            sheet.cell(row=row_number, column=1, value=item["global_id"])
            filling_level = item["filling_level"]
            sheet.cell(
                row=row_number,
                column=2,
                value=(
                    float(filling_level) / 100
                    if filling_level is not None
                    else "—"
                ),
            )
            sheet.cell(
                row=row_number,
                column=3,
                value=item["manager_name"] or "Не назначен",
            )

            if include_recommendations:
                for column, key in enumerate(
                    (
                        "green_recommendations",
                        "yellow_recommendations",
                        "red_recommendations",
                    ),
                    start=4,
                ):
                    cell = sheet.cell(
                        row=row_number,
                        column=column,
                        value=item[key],
                    )
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True,
                    )
            if filling_level is not None:
                sheet.cell(row=row_number, column=2).number_format = "0.0%"
            row_number += 1

        return row_number + 1

    next_row = append_section("Лучшие 10 площадок", top_ten, 1)
    next_row = append_section(
        "Площадки с заполнением ниже 80%",
        below_eighty,
        next_row,
        include_recommendations=True,
    )
    append_section(
        "Карточки без рассчитанного процента",
        without_filling,
        next_row,
    )

    sheet.freeze_panes = "A3"
    _apply_widths(sheet, [16, 30, 36, 48, 48, 48])


def build_monitor_export_xlsx(conn: sqlite3.Connection) -> BytesIO:
    """
    Формирует read-only XLSX-выгрузку мониторинга Инвесткарты РФ.

    Функция только читает сохранённые snapshot и V2-результаты. Она не
    обращается к внешнему API, не пересчитывает оценку и не меняет БД.
    """
    rows = _read_export_rows(conn)
    workbook = Workbook()

    _append_technical_sheet(workbook, rows)
    _append_overview_sheet(workbook, rows)
    _append_manager_summary_sheet(workbook, rows)
    _append_rankings_sheet(workbook, rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
