# ╔══════════════════════════════════════════════════════════════╗
# ║               tools/investmap_export.py                     ║
# ║  Конвертер выгрузки ГИС НСИ (инвестплощадки) → текст        ║
# ║  Формат 1: 1 площадка (атрибут|значение по строкам)         ║
# ║  Формат 2: N площадок (строки=площадки, столбцы=атрибуты)   ║
# ║  Формат 3: 1 площадка ГИС НСИ (шапка на стр.2, 3 столбца)  ║
# ╚══════════════════════════════════════════════════════════════╝

import openpyxl
import io
import re


def _clean(val):
    if val is None:
        return ''
    s = str(val).strip()
    s = re.sub(r'<[^>]+>', '', s)
    return s


def _detect_format(ws):
    """
    Возвращает: 'f3', 'f1', 'f2'

    Формат 3 (ГИС НСИ одна площадка):
      - 3 столбца
      - строка 2 содержит 'Полные наименования атрибутов' или 'Значения атрибутов'
    Формат 1:
      - 2 столбца, много строк
    Формат 2:
      - всё остальное (много столбцов = таблица площадок)
    """
    max_col = ws.max_column
    max_row = ws.max_row

    # Формат 3: 3 столбца И строка 2 похожа на шапку ГИС НСИ
    if max_col <= 4 and max_row >= 3:
        row2 = [_clean(c.value).lower() for c in next(ws.iter_rows(min_row=2, max_row=2))]
        row2_text = ' '.join(row2)
        if 'атрибут' in row2_text or 'значени' in row2_text:
            return 'f3'

    # Формат 1: 2 столбца
    if max_col <= 2 and max_row >= 3:
        return 'f1'

    return 'f2'


def parse_format1(ws):
    """
    Формат 1: 1 площадка.
    Каждая строка: (название атрибута, значение).
    """
    lines = []
    for row in ws.iter_rows(min_row=1):
        cells = [_clean(c.value) for c in row]
        attr = cells[0] if len(cells) > 0 else ''
        val  = cells[1] if len(cells) > 1 else ''
        if not attr:
            continue
        val_out = val if val else 'ПУСТО'
        lines.append(f"{attr} → {val_out}")
    return lines


def parse_format3(ws):
    """
    Формат 3: 1 площадка из ГИС НСИ.
    Строка 1 — заголовок каталога (пропускаем).
    Строка 2 — шапка: [Полные наименования атрибутов, Описание атрибута, Значения атрибутов].
    Строки 3+ — данные: col A = атрибут, col B = описание, col C = значение.
    """
    lines = []
    for row in ws.iter_rows(min_row=3):
        cells = [_clean(c.value) for c in row]
        attr = cells[0] if len(cells) > 0 else ''
        val  = cells[2] if len(cells) > 2 else ''
        if not attr:
            continue
        # Пропускаем служебные/технические поля (архивные, координаты и т.п.)
        val_out = val if val else 'ПУСТО'
        lines.append(f"{attr} → {val_out}")
    return lines


def parse_format2(ws):
    """
    Формат 2: N площадок.
    Строка 1 = заголовки атрибутов.
    Строки 2+ = данные площадок.
    """
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(_clean(cell.value))

    blocks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cells = [_clean(c.value) for c in row]
        if not any(cells):
            continue
        lines = [f"=== ПЛОЩАДКА {row_idx - 1} ==="]
        for h, v in zip(headers, cells):
            if not h:
                continue
            val_out = v if v else 'ПУСТО'
            lines.append(f"{h} → {val_out}")
        blocks.append(lines)
    return blocks


def convert_excel_to_text(file_bytes):
    """
    Основная функция.
    Принимает bytes файла .xlsx.
    Возвращает dict:
      {
        'format': 1, 2 или 3,
        'count': количество площадок,
        'text': итоговый текст для вставки в чат,
        'error': None или строка ошибки
      }
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        fmt = _detect_format(ws)

        if fmt == 'f3':
            lines = parse_format3(ws)
            text = '\n'.join(lines)
            return {
                'format': 3,
                'count': 1,
                'text': text,
                'error': None
            }

        elif fmt == 'f1':
            lines = parse_format1(ws)
            text = '\n'.join(lines)
            return {
                'format': 1,
                'count': 1,
                'text': text,
                'error': None
            }

        else:
            blocks = parse_format2(ws)
            text = '\n\n'.join(['\n'.join(b) for b in blocks])
            return {
                'format': 2,
                'count': len(blocks),
                'text': text,
                'error': None
            }

    except Exception as e:
        return {
            'format': None,
            'count': 0,
            'text': '',
            'error': str(e)
        }
