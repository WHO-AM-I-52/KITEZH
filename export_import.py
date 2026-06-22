# ╔═════════════════════════════════════════════════════════════════════════════╗
# ║                       export_import.py                                       ║
# ║  Импорт обновлённого Excel и валидация выгрузки площадок ГИС НСИ.            ║
# ║  Выделено из export_routes.py (декомпозиция, refactor/structure).            ║
# ║  Функции принимают уже открытую сессию пользователя и возвращают dict-       ║
# ║  результаты; HTTP-обёртки (request/jsonify) остаются в export_routes.py.     ║
# ╚═════════════════════════════════════════════════════════════════════════════╝

from datetime import datetime
import openpyxl

from db import get_db
from core.activity_log import log_action
from export_helpers import (
    STATUS_IMPORT_MAP, REQUIRED_FOR_CREATE,
    STUB_PAYMENT_MAX, VRI_INCOMPATIBLE_WITH_AGRI,
    PRODUCTION_ACTIVITY_KEYWORDS, ROAD_KEYWORDS_IN_TKO,
    _apply_cell_value, _gen_request_number,
)


def process_import_full(file, overwrite: bool, user_id: int) -> tuple:
    """Импорт обновлённого Excel базы обращений.

    Возвращает (payload_dict, http_status). payload_dict готов к jsonify.
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return {'error': f'Ошибка чтения файла: {e}'}, 400

    headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # v3.8: заголовки Excel обновлены под _min/_max
    COL_MAP = {
        '№ обращения':          'request_number',
        'Дата обращения':        'request_date',
        'Полное наименование':   'applicant_full_name',
        'Краткое наименование':  'applicant_short_name',
        'ИНН':                   'applicant_inn',
        'Название проекта':      'project_name',
        'Контактное лицо':       'contact_person',
        'Телефон':               'contact_phone',
        'E-mail':                'contact_email',
        'Инвестиции (млн руб.)': 'investment_total',
        'Рабочих мест':          'jobs_total',
        'Площадь от (га)':       'site_area_ha_min',
        'Площадь до (га)':       'site_area_ha_max',
        'Застройка от (м²)':     'site_build_area_m2_min',
        'Застройка до (м²)':     'site_build_area_m2_max',
        'Районы':                'preferred_districts',
        'Источник':              'source_type',
        'Дата обратной связи':   'feedback_date',
        'Входящий номер':        'incoming_number',
        'Дата ответа':           'answer_date',
        'Способ ответа':         'answer_method',
        'Примечания к ответу':   'answer_notes',
        'Доп. информация':       'additional_info',
    }
    FK_MAP = {
        'Предмет обращения': ('subject_type_id', 'subject_types'),
        'Итоги работы':      ('result_type_id',  'result_types'),
        'Ответственный':     ('assigned_to',     'users'),
    }
    STATUS_COL = 'Статус'

    try:
        id_idx = headers.index('ID (не менять)')
    except ValueError:
        return {'error': 'Колонка «ID (не менять)» не найдена. Используйте файл из «Скачать базу»'}, 400

    status_idx = headers.index(STATUS_COL) if STATUS_COL in headers else None

    conn = get_db()

    updated        = 0
    created        = 0
    skipped        = 0
    status_changed = 0
    errors         = []
    duplicates     = []
    created_ids    = []
    now            = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        subjects  = {r['name']: r['id'] for r in conn.execute('SELECT id,name FROM subject_types').fetchall()}
        results   = {r['name']: r['id'] for r in conn.execute('SELECT id,name FROM result_types').fetchall()}
        users_map = {r['full_name']: r['id'] for r in conn.execute('SELECT id,full_name FROM users').fetchall()}
        fk_lookup = {
            'subject_type_id': subjects,
            'result_type_id':  results,
            'assigned_to':     users_map,
        }

        for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_id = row[id_idx]

            # ── читаем статус ──────────────────────────────────────────────────
            row_status = None
            if status_idx is not None:
                raw_status = row[status_idx]
                if raw_status and str(raw_status).strip():
                    row_status = STATUS_IMPORT_MAP.get(str(raw_status).strip())

            # ── новая строка без ID ─────────────────────────────────────────────
            if not raw_id:
                new_vals = {}
                for ci, header in enumerate(headers):
                    if ci == id_idx:
                        continue
                    if status_idx is not None and ci == status_idx:
                        continue
                    cell_val = row[ci]
                    if header in COL_MAP:
                        field = COL_MAP[header]
                        if cell_val is None or str(cell_val).strip() == '':
                            continue
                        val, ok = _apply_cell_value(field, cell_val, f'Строка {excel_row_num}', errors)
                        if ok and val is not None:
                            new_vals[field] = val
                    elif header in FK_MAP:
                        field, _ = FK_MAP[header]
                        if cell_val is not None and str(cell_val).strip():
                            name = str(cell_val).strip()
                            fk_id = fk_lookup[field].get(name)
                            if fk_id is None:
                                errors.append(f'Строка {excel_row_num}: «{name}» не найдено в справочнике «{header}»')
                            else:
                                new_vals[field] = fk_id

                new_vals['status'] = row_status or 'registered'

                # 3В-2: валидация обязательных полей ───────────────────────────
                missing = [f for f in REQUIRED_FOR_CREATE if not new_vals.get(f)]
                if missing:
                    errors.append(
                        f'Строка {excel_row_num}: пропущена — не заполнены обязательные поля: '
                        + ', '.join(f'«{f}»' for f in missing)
                    )
                    skipped += 1
                    continue

                # ── дедупликация ───────────────────────────────────────────────
                existing_dup = None
                match_by     = None
                inn   = new_vals.get('applicant_inn', '') or ''
                proj  = new_vals.get('project_name', '') or ''
                aname = new_vals.get('applicant_full_name', '') or ''
                rdate = new_vals.get('request_date', '') or ''

                if inn and proj:
                    existing_dup = conn.execute(
                        'SELECT id, status FROM requests WHERE applicant_inn=? AND project_name=?',
                        (inn, proj)
                    ).fetchone()
                    match_by = 'ИНН+проект'
                elif aname and rdate:
                    existing_dup = conn.execute(
                        'SELECT id, status FROM requests WHERE applicant_full_name=? AND request_date=?',
                        (aname, rdate)
                    ).fetchone()
                    match_by = 'наименование+дата'

                if existing_dup:
                    dup_id = existing_dup['id']
                    upd = {k: v for k, v in new_vals.items() if k != 'status'}
                    status_upd = None
                    if row_status and overwrite and row_status != existing_dup['status']:
                        upd['status'] = row_status
                        status_upd = row_status
                    if upd:
                        set_cl = ', '.join(f'{k}=?' for k in upd)
                        conn.execute(
                            f'UPDATE requests SET {set_cl}, updated_at=?, updated_by=? WHERE id=?',
                            list(upd.values()) + [now, user_id, dup_id]
                        )
                        log_action(conn, user_id, 'import_xlsx_dedup', dup_id,
                                   f'Импорт Excel: дедупликация по {match_by} (ИНН={inn or aname})')
                        updated += 1
                        if status_upd:
                            status_changed += 1
                    else:
                        skipped += 1
                    duplicates.append({
                        'existing_id': dup_id,
                        'match_by':    match_by,
                        'inn':         inn or None,
                        'name':        aname or None,
                        'project':     proj or None,
                    })
                else:
                    cols_ins = ', '.join(new_vals.keys()) + ', created_by, created_at, updated_at'
                    ph_ins   = ', '.join(['?'] * len(new_vals)) + ', ?, ?, ?'
                    ins_vals = list(new_vals.values()) + [user_id, now, now]
                    cursor   = conn.execute(
                        f'INSERT INTO requests ({cols_ins}) VALUES ({ph_ins})', ins_vals
                    )
                    new_id = cursor.lastrowid

                    # 3В-3: авто-генерация request_number если не задан ─────────
                    if not new_vals.get('request_number'):
                        auto_num = _gen_request_number(new_id)
                        conn.execute(
                            'UPDATE requests SET request_number=? WHERE id=?',
                            (auto_num, new_id)
                        )

                    log_action(conn, user_id, 'import_xlsx_create', new_id,
                               'Импорт Excel: создано новое обращение')
                    created_ids.append(new_id)
                    created += 1
                continue

            # ── строка с ID → обновляем ─────────────────────────────────────────────
            try:
                rid = int(raw_id)
            except (ValueError, TypeError):
                errors.append(f'Строка {excel_row_num}: невалидный ID: {raw_id}')
                continue

            existing = conn.execute('SELECT * FROM requests WHERE id=?', (rid,)).fetchone()
            if not existing:
                errors.append(f'Строка {excel_row_num} (ID {rid}): обращение не найдено в базе')
                continue

            updates = {}
            status_will_change = False

            if row_status and row_status != existing['status']:
                if overwrite or not existing['status']:
                    updates['status'] = row_status
                    status_will_change = True

            row_label = f'Строка {excel_row_num} (ID {rid})'

            for ci, header in enumerate(headers):
                if ci == id_idx:
                    continue
                if status_idx is not None and ci == status_idx:
                    continue
                cell_val = row[ci]

                if header in COL_MAP:
                    field = COL_MAP[header]
                    if cell_val is None or str(cell_val).strip() == '':
                        continue
                    val, ok = _apply_cell_value(field, cell_val, row_label, errors)
                    if not ok or val is None:
                        continue
                    if not overwrite and existing[field] not in (None, ''):
                        continue
                    updates[field] = val

                elif header in FK_MAP:
                    field, _ = FK_MAP[header]
                    if cell_val is None or str(cell_val).strip() == '':
                        continue
                    name = str(cell_val).strip()
                    fk_id = fk_lookup[field].get(name)
                    if fk_id is None:
                        errors.append(f'{row_label}: «{name}» не найдено в справочнике «{header}»')
                        continue
                    if not overwrite and existing[field] not in (None, ''):
                        continue
                    updates[field] = fk_id

            if not updates:
                skipped += 1
                continue

            set_clause = ', '.join(f'{k}=?' for k in updates)
            vals = list(updates.values()) + [now, user_id, rid]
            conn.execute(
                f'UPDATE requests SET {set_clause}, updated_at=?, updated_by=? WHERE id=?',
                vals
            )
            log_action(conn, user_id, 'import_xlsx', rid,
                       f'Импорт Excel: обновлены поля: {", ".join(updates.keys())}')
            updated += 1
            if status_will_change:
                status_changed += 1

        conn.commit()
    finally:
        conn.close()

    return {
        'updated':        updated,
        'created':        created,
        'skipped':        skipped,
        'status_changed': status_changed,
        'duplicates':     duplicates,
        'created_ids':    created_ids,
        'errors':         errors,
    }, 200


# ─── ВАЛИДАЦИЯ ПЛОЩАДОК ГИС НСИ ─────────────────────────────────────────────
# Используется при импорте выгрузки из ГИС НСИ Нижегородской области.
# Каждая функция принимает словарь записи и возвращает список проблем.

def _parse_coords_point(coords_str: str) -> tuple:
    """БАГ-2: разбирает 'lon,lat' или 'lat,lon' → (lat_wgs84, lon_wgs84).
    ГИС НСИ экспортирует в формате 'longitude,latitude'.
    Возвращает (None, None) если разобрать не удалось.
    """
    if not coords_str or not coords_str.strip():
        return None, None
    s = coords_str.strip().replace(' ', '')
    # Пробуем разделители: запятая, точка с запятой, пробел
    for sep in (',', ';'):
        parts = s.split(sep)
        if len(parts) == 2:
            try:
                first  = float(parts[0].replace(',', '.'))
                second = float(parts[1].replace(',', '.'))
                # ГИС НСИ: первое = долгота (36..45 для НО), второе = широта (54..58)
                if 36.0 <= first <= 50.0 and 50.0 <= second <= 60.0:
                    return second, first   # (lat, lon)
                # Обратный порядок
                if 50.0 <= first <= 60.0 and 36.0 <= second <= 50.0:
                    return first, second   # (lat, lon)
            except ValueError:
                continue
    return None, None


def validate_site_record(rec: dict) -> dict:
    """Валидирует одну запись площадки ГИС НСИ.

    Args:
        rec: словарь с полями записи (ключи = заголовки столбцов выгрузки).

    Returns:
        dict с ключами:
            'errors'   — список критических ошибок (блокируют сохранение)
            'warnings' — список предупреждений (сохраняем, но сигнализируем)
            'fixes'    — словарь автоисправлений {поле: новое_значение}
    """
    errors   = []
    warnings = []
    fixes    = {}

    name = rec.get('Название площадки', '') or ''

    # ── БАГ-14: название со строчной буквы → автоисправление ──────────────
    if name and name[0].islower():
        fixed_name = name[0].upper() + name[1:]
        fixes['Название площадки'] = fixed_name
        warnings.append(
            f'БАГ-14: название начинается со строчной буквы — исправлено: «{fixed_name}»'
        )

    # ── БАГ-2: координаты — авторазбор точки → Широта/Долгота WGS-84 ──────
    coords_pt  = rec.get('Координаты (точка)', '') or ''
    lat_filled = rec.get('Широта WGS-84', '') or ''
    lon_filled = rec.get('Долгота WGS-84', '') or ''

    if coords_pt and not str(lat_filled).strip() and not str(lon_filled).strip():
        lat, lon = _parse_coords_point(coords_pt)
        if lat is not None and lon is not None:
            fixes['Широта WGS-84']  = lat
            fixes['Долгота WGS-84'] = lon
            warnings.append(
                f'БАГ-2: Широта/Долгота WGS-84 пусты — авторазобрано из «{coords_pt}»: '
                f'lat={lat}, lon={lon}'
            )
        else:
            warnings.append(
                f'БАГ-2: не удалось разобрать координаты из «{coords_pt}» — '
                f'заполните Широту/Долготу WGS-84 вручную'
            )

    # ── БАГ-6: плата за подключение — заглушка ──────────────────────────────
    for pay_field in (
        'Плата за подключение к электросетям, руб.',
        'Плата за подключение к газу, руб.',
        'Плата за подключение к воде, руб.',
        'Плата за подключение к теплу, руб.',
        'Плата за подключение к канализации, руб.',
    ):
        pay_val = rec.get(pay_field)
        if pay_val is not None and str(pay_val).strip() not in ('', '—', '-'):
            try:
                pay_num = float(str(pay_val).replace(',', '.').replace('\xa0', '').replace(' ', ''))
                if 0 < pay_num <= STUB_PAYMENT_MAX:
                    warnings.append(
                        f'БАГ-6: «{pay_field}» = {pay_val} руб. — похоже на заглушку '
                        f'(значение ≤ {STUB_PAYMENT_MAX} руб.), уточните реальную плату'
                    )
            except (ValueError, TypeError):
                pass

    # ── БАГ-7: форма сделки есть, а стоимость пустая ─────────────────────────
    deal_form  = rec.get('Форма сделки', '') or ''
    site_price = rec.get('Стоимость объекта, руб.', '') or ''
    if str(deal_form).strip() and not str(site_price).strip():
        warnings.append(
            f'БАГ-7: форма сделки «{deal_form}» указана, но поле «Стоимость объекта» пустое — '
            f'заполните стоимость или уточните условия сделки'
        )

    # ── БАГ-8: категория с/х + ВРИ несовместимые ─────────────────────────────
    catland = rec.get('Категория земель', '') or ''
    vri     = rec.get('Вид разрешённого использования', '') or ''
    if 'сельскохоз' in catland.lower():
        for incompatible_vri in VRI_INCOMPATIBLE_WITH_AGRI:
            if incompatible_vri.lower() in vri.lower():
                errors.append(
                    f'БАГ-8: категория земель «{catland}» несовместима с ВРИ «{vri}» — '
                    f'необходим перевод категории или изменение ВРИ'
                )
                break

    # ── БАГ-10: текст про дорогу в поле ТКО ──────────────────────────────────
    tko_field = rec.get('Объекты ТКО Иные характеристики', '') or ''
    tko_lower = tko_field.lower()
    if any(kw in tko_lower for kw in ROAD_KEYWORDS_IN_TKO) and len(tko_lower) > 5:
        warnings.append(
            f'БАГ-10: в поле «Объекты ТКО Иные характеристики» обнаружено описание дороги/подъезда: '
            f'«{tko_field[:120]}...» — перенесите в поле «Подъездные пути»'
        )

    # ── БАГ-13: ВРИ = Коммунальное, но виды деятельности производственные ──
    activities = rec.get('Виды экономической деятельности', '') or ''
    act_lower  = activities.lower()
    if 'коммунальн' in vri.lower():
        matched = [kw for kw in PRODUCTION_ACTIVITY_KEYWORDS if kw in act_lower]
        if matched:
            warnings.append(
                f'БАГ-13: ВРИ «{vri}» (коммунальное), но среди видов деятельности найдены '
                f'производственные ключевые слова: {", ".join(matched)} — '
                f'проверьте корректность ВРИ'
            )

    return {'errors': errors, 'warnings': warnings, 'fixes': fixes}


def process_import_sites(file, dry_run: bool, user_id: int) -> tuple:
    """Импорт выгрузки площадок ГИС НСИ с валидацией качества данных.

    Возвращает (payload_dict, http_status). payload_dict готов к jsonify.
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return {'error': f'Ошибка чтения файла: {e}'}, 400

    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    headers   = [str(v).strip() if v else '' for v in next(rows_iter)]

    total       = 0
    passed      = 0
    blocked     = 0
    fixes_count = 0
    report      = []

    for row_num, row_vals in enumerate(rows_iter, start=2):
        total += 1
        rec = {headers[i]: (row_vals[i] if i < len(row_vals) else None)
               for i in range(len(headers))}

        result = validate_site_record(rec)

        if result['fixes']:
            fixes_count += len(result['fixes'])

        if result['errors']:
            blocked += 1
        else:
            passed += 1

        name = rec.get('Название площадки') or rec.get('global_id') or f'строка {row_num}'
        report.append({
            'row':      row_num,
            'name':     str(name)[:80],
            'errors':   result['errors'],
            'warnings': result['warnings'],
            'fixes':    result['fixes'],
        })

    log_action(
        get_db(), user_id, 'import_sites_validate',
        detail=(
            f'Валидация площадок ГИС НСИ: всего={total}, прошло={passed}, '
            f'заблокировано={blocked}, автоисправлений={fixes_count}, dry_run={dry_run}'
        )
    )

    return {
        'total':       total,
        'passed':      passed,
        'blocked':     blocked,
        'fixes_count': fixes_count,
        'dry_run':     dry_run,
        'report':      report,
    }, 200
