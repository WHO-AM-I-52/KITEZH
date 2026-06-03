# ╔═════════════════════════════════════════════════════════════════════════════╗
# ║                       export_routes.py                                       ║
# ║  v3.4: 3В-2 валидация обяз. полей + 3В-3 авто-номер ЗУ при импорте          ║
# ╚═════════════════════════════════════════════════════════════════════════════╝

from flask import Blueprint, request, send_file, jsonify, session
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

from db import get_db, REPORTS_DIR
from auth_utils import login_required, get_user_perm
from activity_log import log_action

report_bp = Blueprint('report', __name__)

# Датовые поля — нормализуем через _parse_date_for_db
DATE_FIELDS = {'request_date', 'answer_date', 'feedback_date'}

# Числовые поля — нормализуем через _parse_numeric_for_db
NUMERIC_FIELDS = {'investment_total', 'jobs_total', 'site_area_ha', 'site_build_area_m2'}

# Обязательные поля для создания новой записи при импорте (3В-2)
REQUIRED_FOR_CREATE = ('applicant_full_name', 'request_date')


# ─── ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ────────────────────────────────────────────────────────

def _short_fio(full_name: str) -> str:
    if not full_name:
        return '—'
    parts = full_name.strip().split()
    if len(parts) >= 3:
        return f"{parts[0]} {parts[1][0].upper()}.{parts[2][0].upper()}."
    if len(parts) == 2:
        return f"{parts[0]} {parts[1][0].upper()}."
    return full_name


def _std_border(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _hex_to_argb(hex_color: str) -> str:
    c = hex_color.lstrip('#').upper() if hex_color else ''
    if len(c) == 6:
        return 'FF' + c
    if len(c) == 8:
        return c
    return ''


def _fmt_date(iso) -> str:
    """ISO-дата или datetime → DD.MM.YYYY для вывода."""
    if not iso:
        return '—'
    if isinstance(iso, (date, datetime)):
        return iso.strftime('%d.%m.%Y')
    s = str(iso).strip()[:10]
    try:
        return datetime.strptime(s, '%Y-%m-%d').strftime('%d.%m.%Y')
    except Exception:
        return s or '—'


def _parse_date_for_db(val) -> str | None:
    """Любое представление даты из Excel → YYYY-MM-DD или None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)):
        serial = int(val)
        if 1 <= serial <= 2958465:
            try:
                return (datetime(1899, 12, 30) + timedelta(days=serial)).strftime('%Y-%m-%d')
            except Exception:
                return None
        return None
    s = str(val).strip()
    if not s or s.lower() in ('none', 'null', '—', '-'):
        return None
    FMTS = [
        '%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M', '%-d.%-m.%Y %H:%M:%S',
        '%d.%m.%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y.%m.%d',
    ]
    for fmt in FMTS:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    if len(s) >= 10 and s[4] == '-':
        try:
            return datetime.strptime(s[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None


def _is_empty_cell(val) -> bool:
    """True если ячейка пустая / заполнитель — не нужно писать ошибку."""
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in ('', 'none', 'null', '—', '-')


def _parse_numeric_for_db(val, field: str) -> tuple:
    """Значение из Excel → число для БД или (None, сообщение об ошибке)."""
    if val is None:
        return None, None
    if isinstance(val, bool):
        return None, f'поле «{field}»: булево значение не является числом'
    if isinstance(val, (int, float)):
        if field in ('jobs_total', 'site_build_area_m2'):
            return int(round(val)), None
        return val, None
    s = str(val).strip()
    if not s or s.lower() in ('—', '-', 'none', 'null', 'н/д', 'нет'):
        return None, None
    s = re.sub(r'[^\d\s,.].*$', '', s).strip()
    s = re.sub(r'[\s\xa0]+', '', s)
    s = s.replace(',', '.')
    parts = s.split('.')
    if len(parts) > 2:
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    if not s:
        return None, None
    try:
        num = float(s)
        if field in ('jobs_total', 'site_build_area_m2'):
            return int(round(num)), None
        return num, None
    except ValueError:
        return None, f'поле «{field}»: не удалось распознать число из {val!r}'


def _mln_to_mld(val) -> str:
    if val is None or val == '':
        return '—'
    try:
        mld = float(str(val).replace(',', '.')) / 1000
        return f"{mld:.3f}".rstrip('0').rstrip('.')
    except (ValueError, TypeError):
        return str(val)


def _contact_cell(person: str, phone: str, email: str) -> str:
    parts = []
    if person and person.strip():
        parts.append(f"Ф.И.О. {person.strip()}")
    if phone and phone.strip():
        parts.append(f"Тел: {phone.strip()}")
    if email and email.strip():
        parts.append(email.strip())
    return '\n'.join(parts) if parts else '—'


def _apply_cell_value(field: str, cell_val, row_label: str, errors: list) -> tuple:
    """Универсальный конвертер значения ячейки → (значение_для_бд, успех)."""
    if field in DATE_FIELDS:
        # Пустая ячейка — не ошибка, просто пропускаем
        if _is_empty_cell(cell_val):
            return None, True
        parsed = _parse_date_for_db(cell_val)
        if parsed is None:
            errors.append(f'{row_label}: не удалось распознать дату в поле «{field}»: {cell_val!r}')
            return None, False
        return parsed, True
    if field in NUMERIC_FIELDS:
        num, err_msg = _parse_numeric_for_db(cell_val, field)
        if err_msg:
            errors.append(f'{row_label}: {err_msg}')
            return None, False
        return num, True
    return str(cell_val).strip(), True


def _gen_request_number(new_id: int) -> str:
    """3В-3: авто-генерация номера обращения ЗУ-{ID}-{ГГ}."""
    yy = datetime.now().strftime('%y')
    return f'ЗУ-{new_id}-{yy}'


# ─── СТАНДАРТНАЯ ВЫГРУЗКА ─────────────────────────────────────────────────────────────────────

@report_bp.route('/report')
@login_required
def report():
    df = request.args.get('date_from', '')
    dt = request.args.get('date_to', '')
    sf = request.args.get('status', '')

    conn = get_db()
    try:
        q = ("SELECT r.*,u.full_name as employee_name,ass.full_name as assigned_name "
             "FROM requests r "
             "LEFT JOIN users u   ON r.created_by=u.id "
             "LEFT JOIN users ass ON r.assigned_to=ass.id "
             "WHERE 1=1")
        p = []
        if df:
            q += ' AND r.request_date>=?'; p.append(df)
        if dt:
            q += ' AND r.request_date<=?'; p.append(dt)
        if sf:
            q += ' AND r.status=?'; p.append(sf)

        rows = conn.execute(q + ' ORDER BY r.request_date', p).fetchall()

        sm = {
            'draft':    'Черновик',
            'review':   'На проверке',
            'accepted': 'Принято в работу',
            'answered': 'Ответ направлен',
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Отчёт'

        hfill = PatternFill("solid", fgColor="1B5E7B")
        hfont = Font(bold=True, color="FFFFFF", size=10)
        alt   = PatternFill("solid", fgColor="EAF4FB")
        br    = _std_border()

        ws.merge_cells('A1:Q1')
        per = f" за период {_fmt_date(df)}–{_fmt_date(dt)}" if (df or dt) else ""
        ws['A1'].value     = f"Обращения на подбор земельных участков{per}"
        ws['A1'].font      = Font(bold=True, size=13, color="1B5E7B")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 26

        ws.merge_cells('A2:Q2')
        ws['A2'].value = (
            f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}  "
            f"Всего: {len(rows)}"
        )
        ws['A2'].font      = Font(italic=True, size=9, color="888888")
        ws['A2'].alignment = Alignment(horizontal='center')

        hdrs = [
            '№ обращения', 'Дата', 'Статус', 'Источник', 'Заявитель', 'Название проекта',
            'Контактное лицо', 'Телефон', 'E-mail', 'Инвестиции (млн)',
            'Рабочих мест', 'Площадь (га)', 'Застройка (м²)', 'Право пользования',
            'Районы', 'Ответственный', 'Дата ответа',
        ]
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.fill = hfill; c.font = hfont; c.border = br
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[3].height = 38

        for ri, r in enumerate(rows, 4):
            fill = alt if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            vals = [
                r['request_number'] or '—',
                _fmt_date(r['request_date']),
                sm.get(r['status'], r['status']),
                r['source_type'] or '—',
                r['applicant_short_name'] or r['applicant_full_name'] or '—',
                r['project_name'] or '—',
                r['contact_person'] or '—',
                r['contact_phone'] or '—',
                r['contact_email'] or '—',
                r['investment_total'],
                r['jobs_total'],
                r['site_area_ha'],
                r['site_build_area_m2'],
                r['site_right'] or '—',
                r['preferred_districts'] or '—',
                r['assigned_name'] or r['employee_name'] or '—',
                _fmt_date(r['answer_date']),
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill; c.border = br
                c.alignment = Alignment(vertical='center', wrap_text=True)
            ws.row_dimensions[ri].height = 16

        for ci, w in enumerate(
            [16, 12, 20, 16, 28, 30, 20, 15, 24, 12, 10, 10, 12, 16, 24, 20, 12], 1
        ):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'A4'

        fn = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fp = os.path.join(REPORTS_DIR, fn)
        wb.save(fp)

        log_parts = []
        if df or dt:
            log_parts.append(f"период: {_fmt_date(df)} – {_fmt_date(dt)}")
        if sf:
            log_parts.append(f"статус: {sm.get(sf, sf)}")
        log_parts.append(f"всего {len(rows)} обращ.")
        try:
            log_action(conn, session['user_id'], 'export_report',
                       detail='; '.join(log_parts))
            conn.commit()
        except Exception:
            pass
    finally:
        conn.close()

    return send_file(fp, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── ЕЖЕНЕДЕЛЬНАЯ ВЫГРУЗКА ДЛЯ МИНЭК ───────────────────────────────────────────────────────────────────

@report_bp.route('/report/minek')
@login_required
def report_minek():
    today      = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end   = week_start + timedelta(days=6)

    df = request.args.get('date_from', week_start.isoformat())
    dt = request.args.get('date_to',   week_end.isoformat())
    sf = request.args.get('status', '')

    conn = get_db()
    try:
        q = """
            SELECT
                r.*,
                u.full_name   AS employee_name,
                ass.full_name AS assigned_name,
                st.name       AS subject_type_name,
                rt.name       AS result_type_name,
                rt.color_hex  AS result_color
            FROM requests r
            LEFT JOIN users         u   ON r.created_by      = u.id
            LEFT JOIN users         ass ON r.assigned_to     = ass.id
            LEFT JOIN subject_types st  ON r.subject_type_id = st.id
            LEFT JOIN result_types  rt  ON r.result_type_id  = rt.id
            WHERE r.request_date >= ? AND r.request_date <= ?
        """
        p = [df, dt]
        if sf:
            q += ' AND r.status = ?'; p.append(sf)
        q += ' ORDER BY r.request_date, r.id'

        rows = conn.execute(q, p).fetchall()
        result_types = conn.execute(
            'SELECT id, name, color_hex FROM result_types ORDER BY id'
        ).fetchall()

        HEADER_COLOR = '1B5E7B'
        hfill = PatternFill('solid', fgColor=HEADER_COLOR)
        hfont = Font(bold=True, color='FFFFFF', size=10)
        alt   = PatternFill('solid', fgColor='EAF4FB')
        br    = _std_border()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Заявки'

        NCOLS = 12

        ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
        ws['A1'].value = (
            f"Еженедельный доклад МинЭК: обращения за период "
            f"{_fmt_date(df)} – {_fmt_date(dt)}"
        )
        ws['A1'].font      = Font(bold=True, size=13, color=HEADER_COLOR)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
        ws['A2'].value = (
            f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}   "
            f"Обращений в выборке: {len(rows)}"
        )
        ws['A2'].font      = Font(italic=True, size=9, color='888888')
        ws['A2'].alignment = Alignment(horizontal='center')

        HEADERS = [
            '',
            'Дата обращения',
            'Наименование компании',
            'Наименование проекта',
            'Объем инвестиций,\nмлрд рублей',
            'Рабочие места',
            'Предмет обращения',
            'Дата направления презентации',
            'Дата получения обратной связи',
            'Итоги работы по обращению',
            'Менеджер',
            'Телефон, контактное лицо',
        ]

        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.fill = hfill; c.font = hfont; c.border = br
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[3].height = 40

        for ri, r in enumerate(rows, 4):
            argb = _hex_to_argb(r['result_color']) if r['result_color'] else ''
            if argb:
                rfill = PatternFill('solid', fgColor=argb)
            else:
                rfill = alt if ri % 2 == 0 else PatternFill('solid', fgColor='FFFFFFFF')

            result_val = r['additional_info'] or r['result_type_name'] or '—'

            vals = [
                ri - 3,
                _fmt_date(r['request_date']),
                r['applicant_short_name'] or r['applicant_full_name'] or '—',
                r['project_name'] or '—',
                _mln_to_mld(r['investment_total']),
                r['jobs_total'] or '—',
                r['subject_type_name'] or '—',
                _fmt_date(r['answer_date']),
                _fmt_date(r['feedback_date']),
                result_val,
                _short_fio(r['assigned_name'] or r['employee_name']),
                _contact_cell(
                    r['contact_person'],
                    r['contact_phone'],
                    r['contact_email'],
                ),
            ]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill   = rfill
                c.border = br
                c.alignment = Alignment(
                    vertical='center',
                    wrap_text=True,
                    horizontal='center' if ci == 1 else 'left',
                )
            ws.row_dimensions[ri].height = 30

        col_widths = [5, 13, 28, 35, 12, 12, 22, 16, 16, 30, 16, 32]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = 'B4'

        wl = wb.create_sheet(title='Справочник')
        wl.merge_cells('A1:C1')
        wl['A1'].value     = 'Легенда цветов (итоги работы по обращению)'
        wl['A1'].font      = Font(bold=True, size=12, color=HEADER_COLOR)
        wl['A1'].alignment = Alignment(horizontal='center')
        wl.row_dimensions[1].height = 22

        for ci, h in enumerate(['Цвет', 'Итог', 'Обозначение'], 1):
            c = wl.cell(row=2, column=ci, value=h)
            c.fill = PatternFill('solid', fgColor=HEADER_COLOR)
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.border = _std_border()
            c.alignment = Alignment(horizontal='center', vertical='center')
        wl.row_dimensions[2].height = 20

        if result_types:
            for li, rt in enumerate(result_types, 3):
                argb = _hex_to_argb(rt['color_hex'] or '')
                fill = PatternFill('solid', fgColor=argb) if argb else PatternFill('solid', fgColor='FFFFFFFF')
                ca = wl.cell(row=li, column=1, value='')
                ca.fill = fill; ca.border = _std_border()
                cb = wl.cell(row=li, column=2, value=rt['name'])
                cb.fill = fill
                cb.font = Font(bold=True, size=10); cb.border = _std_border()
                cb.alignment = Alignment(vertical='center')
                cc = wl.cell(row=li, column=3, value=rt['color_hex'])
                cc.font = Font(italic=True, size=9, color='888888'); cc.border = _std_border()
                cc.alignment = Alignment(vertical='center')
                wl.row_dimensions[li].height = 18
        else:
            wl.cell(row=3, column=1,
                    value='Справочник итогов пуст. Добавьте значения в разделе «Справочники».')

        wl.column_dimensions['A'].width = 8
        wl.column_dimensions['B'].width = 36
        wl.column_dimensions['C'].width = 12

        fn = f"minek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fp = os.path.join(REPORTS_DIR, fn)
        wb.save(fp)

        detail = f"период: {_fmt_date(df)} – {_fmt_date(dt)}; всего {len(rows)} обращ."
        try:
            log_action(conn, session['user_id'], 'export_minek', detail=detail)
            conn.commit()
        except Exception:
            pass
    finally:
        conn.close()

    return send_file(fp, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── ПОЛНАЯ ВЫГРУЗКА БАЗЫ ─────────────────────────────────────────────────────────────────────────

@report_bp.route('/export/full')
@login_required
def export_full():
    if not get_user_perm('can_export_full'):
        return jsonify({'error': 'Недостаточно прав: Скачать полную базу (Excel)'}), 403

    conn = get_db()
    try:
        rows = conn.execute("""
            SELECT r.*,
                   ass.full_name AS assigned_name,
                   st.name       AS subject_type_name,
                   rt.name       AS result_type_name
            FROM requests r
            LEFT JOIN users         ass ON r.assigned_to     = ass.id
            LEFT JOIN subject_types st  ON r.subject_type_id = st.id
            LEFT JOIN result_types  rt  ON r.result_type_id  = rt.id
            ORDER BY r.id
        """).fetchall()

        STATUS_EXPORT_MAP = {
            'draft':             'Черновик',
            'registered':        'Зарегистрировано',
            'in_progress':       'В работе',
            'under_review':      'На проверке',
            'ready_to_send':     'Готово к отправке',
            'sent_to_applicant': 'Документы отправлены',
            'closed':            'Закрыто',
        }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'База обращений'

        hfill   = PatternFill("solid", fgColor="1B5E7B")
        id_fill = PatternFill("solid", fgColor="2E4057")
        hfont   = Font(bold=True, color="FFFFFF", size=10)
        br      = _std_border()

        COLS = [
            ('id',                   'ID (не менять)'),
            ('request_number',       '№ обращения'),
            ('request_date',         'Дата обращения'),
            ('status',               'Статус'),
            ('applicant_full_name',  'Полное наименование'),
            ('applicant_short_name', 'Краткое наименование'),
            ('applicant_inn',        'ИНН'),
            ('project_name',         'Название проекта'),
            ('contact_person',       'Контактное лицо'),
            ('contact_phone',        'Телефон'),
            ('contact_email',        'E-mail'),
            ('investment_total',     'Инвестиции (млн руб.)'),
            ('jobs_total',           'Рабочих мест'),
            ('site_area_ha',         'Площадь (га)'),
            ('site_build_area_m2',   'Застройка (м²)'),
            ('preferred_districts',  'Районы'),
            ('source_type',          'Источник'),
            ('assigned_name',        'Ответственный'),
            ('subject_type_name',    'Предмет обращения'),
            ('feedback_date',        'Дата обратной связи'),
            ('result_type_name',     'Итоги работы'),
            ('incoming_number',      'Входящий номер'),
            ('answer_date',          'Дата ответа'),
            ('answer_method',        'Способ ответа'),
            ('answer_notes',         'Примечания к ответу'),
            ('additional_info',      'Доп. информация'),
        ]

        for ci, (field, header) in enumerate(COLS, 1):
            c = ws.cell(row=1, column=ci, value=header)
            c.fill = id_fill if field == 'id' else hfill
            c.font = hfont
            c.border = br
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 36
        ws.freeze_panes = 'A2'

        for ri, r in enumerate(rows, 2):
            row_keys = list(r.keys())
            for ci, (field, _) in enumerate(COLS, 1):
                val = r[field] if field in row_keys else None
                if field == 'status' and val:
                    val = STATUS_EXPORT_MAP.get(val, val)
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = br
                c.alignment = Alignment(vertical='center', wrap_text=(ci == len(COLS)))

        col_widths = [8, 16, 14, 16, 35, 25, 14, 30, 22, 16, 24, 14, 12, 10, 12, 24, 16, 20, 22, 14, 28, 18, 14, 18, 28, 30]
        for ci, w in enumerate(col_widths[:len(COLS)], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        fn = f"sonar_full_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fp = os.path.join(REPORTS_DIR, fn)
        wb.save(fp)

        log_action(conn, session['user_id'], 'export_full',
                   detail=f'Полная выгрузка базы: {len(rows)} обращений')
        conn.commit()
    finally:
        conn.close()

    return send_file(fp, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── ИМПОРТ ОБНОВЛЁННОГО EXCEL ──────────────────────────────────────────────────────────────────────────────────

STATUS_IMPORT_MAP = {
    'Черновик':             'draft',
    'Зарегистрировано':      'registered',
    'В работе':              'in_progress',
    'На проверке':           'under_review',
    'Готово к отправке':     'ready_to_send',
    'Документы отправлены':  'sent_to_applicant',
    'Закрыто':               'closed',
}


@report_bp.route('/import/full', methods=['POST'])
@login_required
def import_full():
    if not get_user_perm('can_import_full'):
        return jsonify({'error': 'Недостаточно прав: Загрузить обновлённый Excel (импорт)'}), 403

    file = request.files.get('import_file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Загрузите файл .xlsx'}), 400

    overwrite = request.form.get('overwrite') == '1'

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'Ошибка чтения файла: {e}'}), 400

    headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    COL_MAP = {
        '№ обращения':          'request_number',
        'Дата обращения':        'request_date',
        'Полное наименование':   'applicant_full_name',
        'Краткое наименование':  'applicant_short_name',
        'ИНН':                   'applicant_inn',
        'Название проекта':      'project_name',
        'Контактное лицо':       'contact_person',
        'Телефон':               'contact_phone',
        'E-mail':                'contact_email',
        'Инвестиции (млн руб.)': 'investment_total',
        'Рабочих мест':          'jobs_total',
        'Площадь (га)':          'site_area_ha',
        'Застройка (м²)':        'site_build_area_m2',
        'Районы':                'preferred_districts',
        'Источник':              'source_type',
        'Дата обратной связи':   'feedback_date',
        'Входящий номер':        'incoming_number',
        'Дата ответа':           'answer_date',
        'Способ ответа':         'answer_method',
        'Примечания к ответу':   'answer_notes',
        'Доп. информация':       'additional_info',
    }
    FK_MAP = {
        'Предмет обращения': ('subject_type_id', 'subject_types'),
        'Итоги работы':      ('result_type_id',  'result_types'),
        'Ответственный':     ('assigned_to',     'users'),
    }
    STATUS_COL = 'Статус'

    try:
        id_idx = headers.index('ID (не менять)')
    except ValueError:
        return jsonify({'error': 'Колонка «ID (не менять)» не найдена. Используйте файл из «Скачать базу»'}), 400

    status_idx = headers.index(STATUS_COL) if STATUS_COL in headers else None

    conn = get_db()

    updated        = 0
    created        = 0
    skipped        = 0
    status_changed = 0
    errors         = []
    duplicates     = []
    created_ids    = []
    now            = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    try:
        subjects  = {r['name']: r['id'] for r in conn.execute('SELECT id,name FROM subject_types').fetchall()}
        results   = {r['name']: r['id'] for r in conn.execute('SELECT id,name FROM result_types').fetchall()}
        users_map = {r['full_name']: r['id'] for r in conn.execute('SELECT id,full_name FROM users').fetchall()}
        fk_lookup = {
            'subject_type_id': subjects,
            'result_type_id':  results,
            'assigned_to':     users_map,
        }

        for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw_id = row[id_idx]

            # ── читаем статус ──────────────────────────────────────────────────
            row_status = None
            if status_idx is not None:
                raw_status = row[status_idx]
                if raw_status and str(raw_status).strip():
                    row_status = STATUS_IMPORT_MAP.get(str(raw_status).strip())

            # ── новая строка без ID ─────────────────────────────────────────────
            if not raw_id:
                new_vals = {}
                for ci, header in enumerate(headers):
                    if ci == id_idx:
                        continue
                    if status_idx is not None and ci == status_idx:
                        continue
                    cell_val = row[ci]
                    if header in COL_MAP:
                        field = COL_MAP[header]
                        if cell_val is None or str(cell_val).strip() == '':
                            continue
                        val, ok = _apply_cell_value(field, cell_val, f'Строка {excel_row_num}', errors)
                        if ok and val is not None:
                            new_vals[field] = val
                    elif header in FK_MAP:
                        field, _ = FK_MAP[header]
                        if cell_val is not None and str(cell_val).strip():
                            name = str(cell_val).strip()
                            fk_id = fk_lookup[field].get(name)
                            if fk_id is None:
                                errors.append(f'Строка {excel_row_num}: «{name}» не найдено в справочнике «{header}»')
                            else:
                                new_vals[field] = fk_id

                new_vals['status'] = row_status or 'registered'

                # 3В-2: валидация обязательных полей ───────────────────────────
                missing = [f for f in REQUIRED_FOR_CREATE if not new_vals.get(f)]
                if missing:
                    errors.append(
                        f'Строка {excel_row_num}: пропущена — не заполнены обязательные поля: '
                        + ', '.join(f'«{f}»' for f in missing)
                    )
                    skipped += 1
                    continue

                # ── дедупликация ───────────────────────────────────────────────
                existing_dup = None
                match_by     = None
                inn   = new_vals.get('applicant_inn', '') or ''
                proj  = new_vals.get('project_name', '') or ''
                aname = new_vals.get('applicant_full_name', '') or ''
                rdate = new_vals.get('request_date', '') or ''

                if inn and proj:
                    existing_dup = conn.execute(
                        'SELECT id, status FROM requests WHERE applicant_inn=? AND project_name=?',
                        (inn, proj)
                    ).fetchone()
                    match_by = 'ИНН+проект'
                elif aname and rdate:
                    existing_dup = conn.execute(
                        'SELECT id, status FROM requests WHERE applicant_full_name=? AND request_date=?',
                        (aname, rdate)
                    ).fetchone()
                    match_by = 'наименование+дата'

                if existing_dup:
                    dup_id = existing_dup['id']
                    upd = {k: v for k, v in new_vals.items() if k != 'status'}
                    status_upd = None
                    if row_status and overwrite and row_status != existing_dup['status']:
                        upd['status'] = row_status
                        status_upd = row_status
                    if upd:
                        set_cl = ', '.join(f'{k}=?' for k in upd)
                        conn.execute(
                            f'UPDATE requests SET {set_cl}, updated_at=?, updated_by=? WHERE id=?',
                            list(upd.values()) + [now, session['user_id'], dup_id]
                        )
                        log_action(conn, session['user_id'], 'import_xlsx_dedup', dup_id,
                                   f'Импорт Excel: дедупликация по {match_by} (ИНН={inn or aname})')
                        updated += 1
                        if status_upd:
                            status_changed += 1
                    else:
                        skipped += 1
                    duplicates.append({
                        'existing_id': dup_id,
                        'match_by':    match_by,
                        'inn':         inn or None,
                        'name':        aname or None,
                        'project':     proj or None,
                    })
                else:
                    cols_ins = ', '.join(new_vals.keys()) + ', created_by, created_at, updated_at'
                    ph_ins   = ', '.join(['?'] * len(new_vals)) + ', ?, ?, ?'
                    ins_vals = list(new_vals.values()) + [session['user_id'], now, now]
                    cursor   = conn.execute(
                        f'INSERT INTO requests ({cols_ins}) VALUES ({ph_ins})', ins_vals
                    )
                    new_id = cursor.lastrowid

                    # 3В-3: авто-генерация request_number если не задан ─────────
                    if not new_vals.get('request_number'):
                        auto_num = _gen_request_number(new_id)
                        conn.execute(
                            'UPDATE requests SET request_number=? WHERE id=?',
                            (auto_num, new_id)
                        )

                    log_action(conn, session['user_id'], 'import_xlsx_create', new_id,
                               'Импорт Excel: создано новое обращение')
                    created_ids.append(new_id)
                    created += 1
                continue

            # ── строка с ID → обновляем ─────────────────────────────────────────────
            try:
                rid = int(raw_id)
            except (ValueError, TypeError):
                errors.append(f'Строка {excel_row_num}: невалидный ID: {raw_id}')
                continue

            existing = conn.execute('SELECT * FROM requests WHERE id=?', (rid,)).fetchone()
            if not existing:
                errors.append(f'Строка {excel_row_num} (ID {rid}): обращение не найдено в базе')
                continue

            updates = {}
            status_will_change = False

            if row_status and row_status != existing['status']:
                if overwrite or not existing['status']:
                    updates['status'] = row_status
                    status_will_change = True

            row_label = f'Строка {excel_row_num} (ID {rid})'

            for ci, header in enumerate(headers):
                if ci == id_idx:
                    continue
                if status_idx is not None and ci == status_idx:
                    continue
                cell_val = row[ci]

                if header in COL_MAP:
                    field = COL_MAP[header]
                    if cell_val is None or str(cell_val).strip() == '':
                        continue
                    val, ok = _apply_cell_value(field, cell_val, row_label, errors)
                    if not ok or val is None:
                        continue
                    if not overwrite and existing[field] not in (None, ''):
                        continue
                    updates[field] = val

                elif header in FK_MAP:
                    field, _ = FK_MAP[header]
                    if cell_val is None or str(cell_val).strip() == '':
                        continue
                    name = str(cell_val).strip()
                    fk_id = fk_lookup[field].get(name)
                    if fk_id is None:
                        errors.append(f'{row_label}: «{name}» не найдено в справочнике «{header}»')
                        continue
                    if not overwrite and existing[field] not in (None, ''):
                        continue
                    updates[field] = fk_id

            if not updates:
                skipped += 1
                continue

            set_clause = ', '.join(f'{k}=?' for k in updates)
            vals = list(updates.values()) + [now, session['user_id'], rid]
            conn.execute(
                f'UPDATE requests SET {set_clause}, updated_at=?, updated_by=? WHERE id=?',
                vals
            )
            log_action(conn, session['user_id'], 'import_xlsx', rid,
                       f'Импорт Excel: обновлены поля: {", ".join(updates.keys())}')
            updated += 1
            if status_will_change:
                status_changed += 1

        conn.commit()
    finally:
        conn.close()

    return jsonify({
        'updated':        updated,
        'created':        created,
        'skipped':        skipped,
        'status_changed': status_changed,
        'duplicates':     duplicates,
        'created_ids':    created_ids,
        'errors':         errors,
    })


# ─── AUTOSAVE / WAL CHECKPOINT ─────────────────────────────────────────────────────────────────────────────────

@report_bp.route('/autosave', methods=['POST'])
@login_required
def autosave():
    try:
        conn = get_db()
        conn.execute('PRAGMA wal_checkpoint(PASSIVE)')
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
